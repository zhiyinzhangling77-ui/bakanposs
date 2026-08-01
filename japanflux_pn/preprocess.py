"""データ読み込み + 5 日周期アノマリ + 欠測処理 (listwise deletion)。

出力はレギュラ 30 分格子上の :class:`PreprocessResult`。「完全格子 + valid mask +
グリッド位置 (step_index)」で保持することで、後段のラグ処理をタイムスタンプ整合で
書け、サイトごとに欠測パターンが違っても同一コードが通る。

読み込み (:func:`load_corevars_hh`) とアノマリ計算 (:func:`compute_anomaly`) を
分離してあるので、実データ無しでもアノマリ・欠測処理を単体テストできる。
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd

from .config import AnalysisConfig, RK_VARS
from .sites import (
    BASE_RH_COL,
    CHINAFLUX_TIME_COLS,
    SiteSpec,
    VPD_FROM_TA_RH,
    get_site,
    resolve_qc_columns,
)


def _vpd_from_ta_rh(ta_c: np.ndarray, rh_pct: np.ndarray) -> np.ndarray:
    """気温[°C]・相対湿度[%]から飽差 VPD[hPa] を導出する。

    BASE 形式 (韓国・中国) は VPD 列を持たないため、飽和水蒸気圧 e_s を
    Magnus-Tetens (Alduchov & Eskridge 1996) で求め、VPD = e_s·(1−RH/100)。
    単位 hPa は JapanFlux2024 の VPD_F と一致する。入力欠測は NaN 伝播。
    """
    es = 6.1094 * np.exp(17.625 * ta_c / (ta_c + 243.04))   # hPa
    vpd = es * (1.0 - rh_pct / 100.0)
    return np.where(np.isfinite(ta_c) & np.isfinite(rh_pct), vpd, np.nan)


@dataclass
class PreprocessResult:
    """前処理の成果物。"""

    anomaly: pd.DataFrame        # レギュラ格子 × RK_VARS、無効時刻は NaN
    valid: pd.Series             # 全 11 変数が揃う時刻 True (bool, 格子と同 index)
    site: str
    year: int
    month: int                   # 代表月 (プール時は先頭月)
    config: AnalysisConfig
    months: list[int] | None = None   # プール対象月 (単月なら [month])

    @property
    def month_label(self) -> str:
        """出力ディレクトリ等に使う月ラベル ("07" or "07-08")。"""
        ms = self.months or [self.month]
        return f"{ms[0]:02d}" if len(ms) == 1 else f"{ms[0]:02d}-{ms[-1]:02d}"

    @property
    def valid_frame(self) -> pd.DataFrame:
        """時間順に詰めた有効サンプルのみの DataFrame (NaN 無し)。"""
        return self.anomaly.loc[self.valid]

    @property
    def step_index(self) -> np.ndarray:
        """有効サンプルのレギュラ格子上の整数位置 k (= (t - t0)/Δt)。"""
        idx = self.anomaly.index
        t0 = idx[0]
        step = pd.Timedelta(minutes=24 * 60 // self.config.steps_per_day)
        k_all = ((idx - t0) / step).astype("int64").to_numpy()
        return k_all[self.valid.to_numpy()]

    @property
    def n_points(self) -> int:
        return int(self.valid.sum())

    def meta(self) -> dict:
        """成果物に添えるメタ情報。"""
        return {
            "site": self.site,
            "year": self.year,
            "month": self.month,
            "months": "+".join(str(m) for m in (self.months or [self.month])),
            "n_points": self.n_points,
            "n_bins": self.config.n_bins,
            "anomaly_window_days": self.config.anomaly_window_days,
            "sig_c": self.config.sig_c,
            "n_surrogates": self.config.n_surrogates,
            "lag_min_h": self.config.lag_hours(self.config.lag_min),
            "lag_max_h": self.config.lag_hours(self.config.lag_max),
        }


# ---------------------------------------------------------------------------
# 5 日周期アノマリ (R&K §3)
# ---------------------------------------------------------------------------
def _forward_window_anomaly(
    values: np.ndarray, steps_per_day: int, window_days: int
) -> np.ndarray:
    """各時刻 t の値から、同時刻帯の直後 window_days 日平均を引く。

    平均は ``{x(t + d·steps_per_day) : d = 0..W-1}`` (t を day0 に含む前方窓)。
    窓 W 点が全て有限のときのみアノマリを確定し、1 点でも欠測なら NaN。
    末尾 (W-1)·steps_per_day サンプルは窓が満たせず NaN になる。
    """
    n = len(values)
    stack = np.full((window_days, n), np.nan)
    for d in range(window_days):
        off = d * steps_per_day
        if off < n:
            stack[d, : n - off] = values[off:]
    full = np.all(np.isfinite(stack), axis=0)
    wmean = np.where(full, np.mean(stack, axis=0), np.nan)
    return values - wmean


def compute_anomaly(
    raw: pd.DataFrame,
    config: AnalysisConfig,
    keep_index: pd.DatetimeIndex | None = None,
) -> tuple[pd.DataFrame, pd.Series]:
    """レギュラ格子上の生 DataFrame から 5 日アノマリと valid mask を作る。

    Parameters
    ----------
    raw:
        DatetimeIndex を持ち RK_VARS 列を含む生値 (欠測は NaN)。格子はレギュラ
        (欠測時刻の行も NaN で存在) であることを前提とする。前方窓のため対象月の
        末尾に window_days 日ぶんのバッファを含めておくと末尾の損失を防げる。
    keep_index:
        アノマリ計算後に残す時刻 (通常は対象月のみ)。None なら全格子を残す。
    """
    sp = config.steps_per_day
    w = config.anomaly_window_days
    anom = pd.DataFrame(index=raw.index, columns=RK_VARS, dtype=float)
    for v in RK_VARS:
        anom[v] = _forward_window_anomaly(raw[v].to_numpy(dtype=float), sp, w)
    if keep_index is not None:
        anom = anom.reindex(keep_index)
    valid = anom.notna().all(axis=1)
    valid.name = "valid"
    return anom, valid


# ---------------------------------------------------------------------------
# 読み込み
# ---------------------------------------------------------------------------
def _regular_grid(start: pd.Timestamp, end: pd.Timestamp, steps_per_day: int) -> pd.DatetimeIndex:
    freq = pd.Timedelta(minutes=24 * 60 // steps_per_day)
    return pd.date_range(start=start, end=end, freq=freq)


def _is_xlsx(path: str | Path) -> bool:
    return str(path).lower().endswith((".xlsx", ".xls"))


def _read_table_header(path: str | Path) -> list[str]:
    """csv/xlsx のヘッダ列名だけを取る（BASE 形式が KoFlux では xlsx で来るため）。"""
    if _is_xlsx(path):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        header = [("" if c is None else str(c)) for c in next(ws.iter_rows(values_only=True))]
        wb.close()
        return header
    return list(pd.read_csv(path, nrows=0).columns)


def _read_table_columns(path: str | Path, want: set[str]) -> pd.DataFrame:
    """csv/xlsx から ``want`` に含まれる列だけを読み込む（TIMESTAMP_START 形式想定）。"""
    if _is_xlsx(path):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        header = [("" if c is None else str(c)) for c in next(it)]
        idx = {h: i for i, h in enumerate(header) if h in want}
        data: dict[str, list] = {h: [] for h in idx}
        for row in it:
            for h, i in idx.items():
                data[h].append(row[i] if i < len(row) else None)
        wb.close()
        return pd.DataFrame(data)
    return pd.read_csv(path, usecols=lambda c: c in want)


def read_corevars_raw(
    path: str | Path, site: SiteSpec, config: AnalysisConfig
) -> pd.DataFrame:
    """COREVARS HH CSV を読み、RK_VARS 列 (実カラム→R&K表記) の生 DataFrame にする。

    TIMESTAMP_START (YYYYMMDDHHMM) を index に採用し、欠測センチネルを NaN 化する。
    ``config.qc_max`` が設定されていれば、対応する ``_QC`` 列が閾値を超える (低品質
    gap-fill) 値を NaN 化する（実測寄りデータでの感度解析用）。
    """
    vmap = site.var_map()
    # 実在する値列のみ (VPD 導出マーカは実列ではないので除く)。
    val_cols = [vmap[v] for v in RK_VARS if not vmap[v].startswith("@")]
    derived_vpd = vmap["VPD"] == VPD_FROM_TA_RH
    header = _read_table_header(path)
    qcmap = (resolve_qc_columns(header, site) if config.qc_max is not None
             else {v: None for v in RK_VARS})
    qc_cols = sorted({c for c in qcmap.values() if c})
    want = set(["TIMESTAMP_START"] + val_cols + qc_cols)
    if derived_vpd:                     # VPD 導出には TA と RH の実列が要る
        want |= {vmap["Ta"], BASE_RH_COL}

    df = _read_table_columns(path, want)
    ts = pd.to_datetime(
        pd.to_numeric(df["TIMESTAMP_START"]).astype("int64").astype(str),
        format="%Y%m%d%H%M")
    df = df.drop(columns=["TIMESTAMP_START"])
    df.index = ts
    df = df.replace(config.na_sentinel, np.nan)

    out = pd.DataFrame(index=df.index)
    for v in RK_VARS:
        col = vmap[v]
        if col == VPD_FROM_TA_RH:       # BASE 形式: TA + RH から VPD を導出
            s = pd.Series(
                _vpd_from_ta_rh(
                    df[vmap["Ta"]].to_numpy(dtype=float),
                    df[BASE_RH_COL].to_numpy(dtype=float),
                ),
                index=df.index,
            )
        else:
            s = df[col]
        qc = qcmap[v]
        if config.qc_max is not None and qc is not None:
            s = s.where(df[qc] <= config.qc_max)   # QC>閾値 (or QC欠測) → NaN
        out[v] = s
    return out[RK_VARS]


def read_chinaflux_raw(
    path: str | Path, site: SiteSpec, config: AnalysisConfig
) -> pd.DataFrame:
    """ChinaFlux の 30 分 HH xlsx を読み、RK_VARS 列の生 DataFrame にする。

    ChinaFlux 特有の扱い:
      - 時刻は Year/Month/Day/hour/min の 5 列から合成 (:data:`CHINAFLUX_TIME_COLS`)。
      - ヘッダ行の直後に「単位行」があるので、Year が整数化できない行はスキップ。
      - 列名は大小が揺れる (ER_DT / ER_dt) ため大小無視で解決する。
      - 欠測は複数のセンチネル (-9999 / -99999 / -6999) を採るため、-6999 以下を NaN 化。
    情報量は単調変換・符号反転・単位に不変なので、NEP=−NEE・mg CO2 単位は変換しない。
    """
    import openpyxl

    vmap = site.var_map()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = [("" if c is None else str(c)) for c in next(it)]
    low = {h.lower(): i for i, h in enumerate(header)}

    def _idx(name: str) -> int:
        if name in header:
            return header.index(name)
        if name.lower() in low:
            return low[name.lower()]
        raise KeyError(f"column {name!r} not in ChinaFlux header of {Path(path).name}")

    t_idx = [_idx(c) for c in CHINAFLUX_TIME_COLS]
    v_idx = {v: _idx(vmap[v]) for v in RK_VARS}

    times: list[pd.Timestamp] = []
    cols: dict[str, list[float]] = {v: [] for v in RK_VARS}
    for row in it:
        try:                                   # 単位行・空行を弾く (Year が整数でない)
            y = int(float(row[t_idx[0]]))
            mo = int(float(row[t_idx[1]])); d = int(float(row[t_idx[2]]))
            hh = int(float(row[t_idx[3]])); mi = int(float(row[t_idx[4]]))
            ts = pd.Timestamp(year=y, month=mo, day=d, hour=hh, minute=mi)
        except (TypeError, ValueError):
            continue
        times.append(ts)
        for v in RK_VARS:
            try:
                cols[v].append(float(row[v_idx[v]]))
            except (TypeError, ValueError):
                cols[v].append(np.nan)
    wb.close()

    out = pd.DataFrame(cols, index=pd.DatetimeIndex(times))
    out = out[~out.index.duplicated(keep="first")].sort_index()
    out = out.mask(out <= -6999.0)             # 複数センチネルをまとめて NaN 化
    return out[RK_VARS]


def read_site_raw(
    path: str | Path, site: SiteSpec, config: AnalysisConfig
) -> pd.DataFrame:
    """サイトの形式に応じて 1 ファイルを読み分けるディスパッチャ。"""
    if site.fmt == "chinaflux":
        return read_chinaflux_raw(path, site, config)
    return read_corevars_raw(path, site, config)   # japanflux / base (csv)


def find_corevars_files(site: SiteSpec) -> list[Path]:
    """サイトの HH ファイル一覧 (ソート済み, 既定は ALLVARS)。無ければ例外。"""
    files = sorted(f for f in Path(site.data_dir).glob(site.glob)
                   if not f.name.startswith("~$"))   # Excel 一時ファイルを除外
    if not files:
        raise FileNotFoundError(
            f"no {site.source} HH file under {site.data_dir!r} "
            f"matching {site.glob!r}"
        )
    return files


def load_raw_all(site: SiteSpec, config: AnalysisConfig) -> pd.DataFrame:
    """サイトの全 COREVARS HH を連結し、RK_VARS 列の生 DataFrame にする。

    複数ファイル (年分割) を時刻でソート・重複除去して 1 本にまとめる。年スキャンで
    繰り返し使うため読み込みを 1 回に集約する用途。
    """
    files = find_corevars_files(site)
    raw_all = pd.concat([read_site_raw(f, site, config) for f in files])
    return raw_all[~raw_all.index.duplicated(keep="first")].sort_index()


def slice_and_anomaly(
    raw_all: pd.DataFrame, year: int, month: int, config: AnalysisConfig
) -> tuple[pd.DataFrame, pd.Series]:
    """連結済み生データから対象月を切り出し、5 日アノマリ + valid mask を作る。

    対象月 + 前方窓バッファ (window_days 日) をレギュラ格子へ張り直し、アノマリ後に
    対象月のみへ絞る。:func:`load_corevars_hh` と年スキャンで共有するコア。
    """
    month_start = pd.Timestamp(year=year, month=month, day=1)
    month_end = month_start + pd.offsets.MonthBegin(1)
    buf_end = month_end + pd.Timedelta(days=config.anomaly_window_days)
    step = pd.Timedelta(minutes=24 * 60 // config.steps_per_day)

    grid = _regular_grid(month_start, buf_end - step, config.steps_per_day)
    raw = raw_all.reindex(grid)
    keep = _regular_grid(month_start, month_end - step, config.steps_per_day)
    return compute_anomaly(raw, config, keep_index=keep)


def slice_span_and_anomaly(
    raw_all: pd.DataFrame, year: int, months: list[int], config: AnalysisConfig
) -> tuple[pd.DataFrame, pd.Series]:
    """連続する複数月を 1 本の連続区間としてプールし、アノマリ + valid mask を作る。

    ``months`` は同一年内の連続月を想定 (例 [7, 8])。先頭月の月初から末尾月の月末
    までを 1 つのレギュラ格子に張り、前方窓バッファを足してアノマリ化する。区間内は
    連続なので、後段のラグ三つ組は ``gap_guard`` で欠測跨ぎだけを避ければよい。
    """
    months = sorted(set(months))
    first_start = pd.Timestamp(year=year, month=months[0], day=1)
    last_start = pd.Timestamp(year=year, month=months[-1], day=1)
    span_end = last_start + pd.offsets.MonthBegin(1)
    buf_end = span_end + pd.Timedelta(days=config.anomaly_window_days)
    step = pd.Timedelta(minutes=24 * 60 // config.steps_per_day)

    grid = _regular_grid(first_start, buf_end - step, config.steps_per_day)
    raw = raw_all.reindex(grid)
    keep = _regular_grid(first_start, span_end - step, config.steps_per_day)
    return compute_anomaly(raw, config, keep_index=keep)


def load_corevars_hh(
    site_code: str,
    year: int,
    month: int | list[int],
    config: AnalysisConfig | None = None,
) -> PreprocessResult:
    """サイト・年・月を指定して COREVARS HH を前処理する。

    ``month`` は単月 (int) または同一年内の連続複数月 (list, 例 [7, 8]) を受ける。
    複数月ならプールして n を稼ぎ、TE 推定器の有限標本 floor を下げる用途。対象区間
    + 前方窓バッファ (window_days 日) の範囲を読み、5 日アノマリ後に対象区間へ絞る。
    """
    config = config or AnalysisConfig()
    site = get_site(site_code)
    raw_all = load_raw_all(site, config)
    months = [month] if isinstance(month, int) else sorted(set(month))
    if len(months) == 1:
        anom, valid = slice_and_anomaly(raw_all, year, months[0], config)
    else:
        anom, valid = slice_span_and_anomaly(raw_all, year, months, config)
    return PreprocessResult(
        anomaly=anom, valid=valid, site=site_code, year=year,
        month=months[0], months=months, config=config,
    )
