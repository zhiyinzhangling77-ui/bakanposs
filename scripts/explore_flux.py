#!/usr/bin/env python3
"""任意のフラックスデータ配布フォルダの構造とヘッダを調べて表示する。

ChinaFlux / KoFlux のような未知フォーマットの中身を確認するための道具。
各ルートについて (1) 直下の一覧, (2) csv ファイル一覧, (3) 代表 csv のヘッダ列と
推定フォーマット・推定解像度 を出す。列名から R&K 11 変数に対応しそうな列も推測する。

    python scripts/explore_flux.py /mnt/hdd/KoFlux /mnt/hdd/ChinaFlux
    python scripts/explore_flux.py /mnt/hdd/KoFlux --max-headers 6
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path


def _list_dir(root: Path, limit: int = 40) -> None:
    print(f"\n[直下] {root}")
    if not root.exists():
        print("  (存在しません)")
        return
    entries = sorted(root.iterdir())
    for p in entries[:limit]:
        kind = "DIR " if p.is_dir() else "file"
        try:
            size = "" if p.is_dir() else f"{p.stat().st_size/1e6:6.1f}MB"
        except OSError:
            size = ""
        print(f"  {kind} {size:>9}  {p.name}")
    if len(entries) > limit:
        print(f"  … 他 {len(entries)-limit} 件")


_TABULAR_EXT = (".csv", ".txt", ".dat", ".tsv")


def _all_files(root: Path, cap: int = 5000) -> list[Path]:
    out = []
    for p in root.rglob("*"):
        if p.is_file():
            out.append(p)
            if len(out) >= cap:
                break
    return out


def _ext_histogram(files: list[Path]) -> None:
    from collections import Counter
    c = Counter((p.suffix.lower() or "(拡張子なし)") for p in files)
    print(f"\n[拡張子の内訳] 総ファイル {len(files)} 件")
    for ext, n in c.most_common(15):
        print(f"  {ext:<14} {n:>5}")


def _sample_leaf_dirs(root: Path, n: int = 2) -> None:
    """代表的な末端フォルダの中身を数個ずつ表示（実データの置き方を見る）。"""
    subdirs = [p for p in sorted(root.rglob("*")) if p.is_dir()]
    shown = 0
    for d in subdirs:
        files = [p for p in sorted(d.iterdir()) if p.is_file()]
        if not files:
            continue
        print(f"\n  [フォルダ] {d}")
        for p in files[:8]:
            try:
                size = f"{p.stat().st_size/1e6:6.1f}MB"
            except OSError:
                size = ""
            print(f"     {size:>9}  {p.name}")
        if len(files) > 8:
            print(f"     … 他 {len(files)-8} 件")
        shown += 1
        if shown >= n:
            break


def _find_csvs(root: Path, limit: int = 30) -> list[Path]:
    """表形式らしきファイル（tabular 拡張子 + xlsx）を集める。"""
    exts = _TABULAR_EXT + (".xlsx", ".xls")
    return sorted(p for p in _all_files(root) if p.suffix.lower() in exts)


def _read_header_rows(path: Path, n_rows: int = 4) -> list[list[str]]:
    """csv/txt/xlsx から先頭数行を「セルのリスト」で返す（区切り自動判定）。"""
    if path.suffix.lower() in (".xlsx", ".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                rows.append(["" if v is None else str(v) for v in row])
                if i + 1 >= n_rows:
                    break
            wb.close()
            return rows
        except Exception as e:  # noqa: BLE001
            return [[f"(xlsx 読めません: {e})"]]
    # テキスト系
    try:
        with open(path, "r", errors="replace") as f:
            lines = [f.readline().rstrip("\n") for _ in range(n_rows)]
    except OSError as e:
        return [[f"(読めません: {e})"]]
    for sep in (",", "\t", ";"):
        if len(lines[0].split(sep)) > 3:
            return [ln.split(sep) for ln in lines if ln]
    return [[lines[0]]]


def _guess_format(header: list[str]) -> str:
    h = set(header)
    joined = " ".join(header)
    if any(c.endswith("_1_1_1") for c in header):
        return "BASE (_1_1_1 位置修飾子; AmeriFlux/ICOS/KoFlux 系)"
    if "TIMESTAMP_START" in h and any("_vUT" in c or "_VUT" in c or "_F" == c[-2:]
                                      for c in header):
        return "FLUXNET2015/JapanFlux 系 (_F / _vUT)"
    if any(k in joined for k in ("NEP_", "ER_DT", "GPP_DT", "_sc", "_avail")):
        return "ChinaFlux 集計系 (NEP_sc/ER_DT 等)"
    return "不明 (要目視)"


def _guess_resolution(header: list[str], sample_path: Path) -> str:
    name = sample_path.name.lower()
    for tag, res in (("hh", "30分(HH)"), ("30", "30分?"), ("hr", "時別"),
                     ("hour", "時別"), ("daily", "日"), ("day", "日"),
                     ("month", "月"), ("year", "年"), ("annual", "年")):
        if tag in name:
            return res
    return "?(ファイル名から不明)"


# R&K 11 変数に対応しそうな列を推測するためのキーワード
RK_HINTS = {
    "Rg(放射)":   ("SW_IN", "RG", "DR", "PAR", "PPFD", "RAD", "SWIN"),
    "Ta(気温)":   ("TA", "TAIR", "T_AIR", "TEMP"),
    "VPD":        ("VPD",),
    "Ts(地温)":   ("TS", "TSOIL", "T_SOIL", "STEMP"),
    "P(降水)":    ("P_", "PREC", "RAIN", "PPT"),
    "θ(土壌水分)": ("SWC", "VWC", "SW1", "SW2", "SOILW", "THETA"),
    "H(顕熱)":    ("H_", "HS", "SH"),
    "LE(潜熱)":   ("LE", "LATENT"),
    "GER/RECO":   ("RECO", "ER_", "RE_", "RESP"),
    "NEE/NEP":    ("NEE", "NEP", "FC", "FCO2"),
    "GEP/GPP":    ("GPP", "GEP"),
}


def _match_rk(header: list[str]) -> None:
    up = [c.upper() for c in header]
    for var, keys in RK_HINTS.items():
        hits = []
        for i, c in enumerate(up):
            if c.startswith("TIMESTAMP") or c in ("TIME", "DATE"):
                continue
            if any(c.startswith(k) for k in keys):   # 前方一致のみ（部分一致のノイズ回避）
                hits.append(header[i])
        shown = ", ".join(hits[:6]) + (" …" if len(hits) > 6 else "")
        print(f"    {var:<12} → {shown or '(該当なし)'}")


def explore(root_str: str, max_headers: int = 5) -> None:
    root = Path(root_str)
    print("\n" + "=" * 70)
    print(f"### {root}")
    print("=" * 70)
    _list_dir(root)

    allf = _all_files(root)
    _ext_histogram(allf)               # ← どんな拡張子で入っているか
    _sample_leaf_dirs(root, n=3)       # ← 末端フォルダの実ファイルを覗く

    csvs = _find_csvs(root)
    print(f"\n[表形式(.csv/.txt/.dat/.tsv/.xlsx) ファイル] {len(csvs)} 件（先頭のみ表示）")
    for p in csvs[:20]:
        try:
            rel = p.relative_to(root)
        except ValueError:
            rel = p
        print(f"  {rel}")

    # 代表ヘッダ: 30分値(HH) と 水田(CRK/paddy/rice) を優先しつつ、名前パターンが
    # 異なるものを数個。
    def _priority(p: Path) -> tuple:
        nm = p.name.upper()
        is_hh = ("HH" in nm) or ("_HH_" in nm)
        is_paddy = any(k in nm for k in ("CRK", "PADDY", "RICE"))
        is_all = "ALL" in nm and "LITE" not in nm
        return (not is_paddy, not is_hh, not is_all, str(p))

    ordered = sorted(csvs, key=_priority)
    seen_keys = set()
    picked = []
    for p in ordered:
        key = (p.parent.parent.name, "HH" in p.name.upper())
        if key not in seen_keys:
            seen_keys.add(key)
            picked.append(p)
        if len(picked) >= max_headers:
            break

    for p in picked:
        print(f"\n--- ヘッダ: {p} ---")
        rows = _read_header_rows(p, n_rows=4)
        cols = rows[0] if rows else []
        print(f"  列数={len(cols)}")
        print(f"  列: {cols[:50]}")
        if len(rows) > 1:
            print(f"  2行目(単位/値?): {rows[1][:50]}")
        print(f"  推定フォーマット: {_guess_format(cols)}")
        print(f"  推定解像度: {_guess_resolution(cols, p)}")
        print("  R&K 11 変数への対応候補:")
        _match_rk(cols)


def main(argv: list[str] | None = None) -> None:
    ap = argparse.ArgumentParser(description="explore flux data folders")
    ap.add_argument("roots", nargs="+", help="調べるフォルダ (複数可)")
    ap.add_argument("--max-headers", type=int, default=5,
                    help="ルートごとに表示する代表ヘッダ数 (既定 5)")
    args = ap.parse_args(argv)
    for r in args.roots:
        explore(r, args.max_headers)


if __name__ == "__main__":
    main(sys.argv[1:] if len(sys.argv) > 1 else None)
